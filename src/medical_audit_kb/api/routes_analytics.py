from __future__ import annotations

import csv
from collections import Counter
from collections.abc import Callable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Annotated, Literal
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from medical_audit_kb.api.app import ApiState, get_api_state, record_operation
from medical_audit_kb.api.auth import (
    Permission,
    require_permission,
    resolve_authenticated_user,
    user_has_permission,
)

router = APIRouter(prefix="/analytics")

MAX_UPLOAD_BYTES = 20 * 1024 * 1024
SUPPORTED_EXTENSIONS = {"csv", "xlsx", "xlsm"}
ColumnType = Literal["数值", "日期", "标识", "文本", "空列"]
AnalysisCase = Literal["audit-data", "dupont"]
CaseStatus = Literal["completed", "needs-input"]


class TableColumnProfile(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str
    type: ColumnType
    empty_count: int
    unique_count: int
    sample_values: list[str]
    audit_hint: str


class TableAnalysisMetric(BaseModel):
    model_config = ConfigDict(extra="forbid")

    key: str
    label: str
    value: float | int | None
    display_value: str
    formula: str | None = None
    status: Literal["available", "unavailable"]


class TableUploadAnalysisResponse(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str
    analysis_case: AnalysisCase
    analysis_case_label: str
    case_status: CaseStatus
    case_metrics: list[TableAnalysisMetric]
    case_findings: list[str]
    size_kb: int
    extension: str
    status: Literal["parsed"]
    sheet_name: str | None
    columns: list[TableColumnProfile]
    row_count: int
    empty_cell_count: int
    duplicate_row_count: int
    message: str
    quality_findings: list[str]
    audit_signals: list[str]
    recommendations: list[str]
    upload_id: str | None = None
    sha256: str | None = None
    retention_status: Literal["retained", "not-configured"] = "not-configured"
    created_at: str | None = None


class AnalyticsUploadHistoryItem(BaseModel):
    model_config = ConfigDict(extra="forbid")

    id: str
    name: str
    analysis_case: AnalysisCase = "audit-data"
    analysis_case_label: str = "审计数据分析"
    extension: str
    size_bytes: int
    size_kb: int
    sha256: str
    storage_path: str = Field(exclude=True)
    sheet_name: str | None
    row_count: int
    column_count: int
    empty_cell_count: int
    duplicate_row_count: int
    status: str
    created_by: str | None
    created_at: str
    retention_status: Literal["retained"]
    audit_signals: list[str]


class AnalyticsUploadHistoryResponse(BaseModel):
    model_config = ConfigDict(extra="forbid")

    items: list[AnalyticsUploadHistoryItem]
    store: dict[str, object]


@router.post("/table-upload", response_model=TableUploadAnalysisResponse)
async def analyze_table_upload(
    file: Annotated[UploadFile, File()],
    state: Annotated[ApiState, Depends(get_api_state)],
    analysis_case: Annotated[AnalysisCase, Form()] = "audit-data",
    x_user_id: Annotated[str | None, Header(alias="X-User-Id")] = None,
    x_role: Annotated[str | None, Header(alias="X-Role")] = None,
) -> TableUploadAnalysisResponse:
    normalized_user_identifier = (x_user_id or "").strip()
    if not normalized_user_identifier or normalized_user_identifier == "anonymous":
        raise HTTPException(status_code=401, detail="X-User-Id header is required")
    user = require_permission(
        state,
        permission=Permission.ANALYZE_DATA,
        x_user_id=normalized_user_identifier,
        x_role=x_role,
        attempted_action="analytics-table-upload",
    )
    file_name = file.filename or "uploaded-table"
    extension = _file_extension(file_name)
    if extension not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=422, detail="unsupported table file extension")

    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="uploaded table file is too large")
    if not content:
        raise HTTPException(status_code=422, detail="uploaded table file is empty")

    sheet_name: str | None = None
    try:
        if extension == "csv":
            rows = _read_csv_rows(content)
        else:
            sheet_name, rows = _read_workbook_rows(content)
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=422,
            detail="table file text encoding is not supported",
        ) from exc
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise HTTPException(status_code=422, detail="workbook file cannot be parsed") from exc

    if not rows:
        raise HTTPException(status_code=422, detail="table file does not contain tabular rows")

    response = _build_response(
        file_name=file_name,
        size_bytes=len(content),
        extension=extension,
        rows=rows,
        sheet_name=sheet_name,
        analysis_case=analysis_case,
    )
    upload_record: dict[str, object] | None = None
    if state.analytics_upload_store is not None:
        upload_record = state.analytics_upload_store.add_upload(
            file_name=file_name,
            extension=extension,
            content=content,
            analysis_summary=_analysis_summary(response),
            created_by=user.user_identifier,
        )
        response = response.model_copy(
            update={
                "upload_id": upload_record["id"],
                "sha256": upload_record["sha256"],
                "retention_status": upload_record["retention_status"],
                "created_at": upload_record["created_at"],
            }
        )
    record_operation(
        state,
        "analytics-table-upload",
        {
            "upload_id": upload_record.get("id") if upload_record else None,
            "file_name": file_name,
            "analysis_case": response.analysis_case,
            "extension": extension,
            "sheet_name": sheet_name,
            "row_count": response.row_count,
            "column_count": len(response.columns),
            "retention_status": response.retention_status,
            "actor": user.user_identifier,
        },
    )
    return response


@router.get("/table-uploads", response_model=AnalyticsUploadHistoryResponse)
def list_table_uploads(
    state: Annotated[ApiState, Depends(get_api_state)],
    limit: Annotated[int, Query(ge=1, le=100)] = 20,
    x_user_id: Annotated[str | None, Header(alias="X-User-Id")] = None,
    x_role: Annotated[str | None, Header(alias="X-Role")] = None,
) -> AnalyticsUploadHistoryResponse:
    normalized_user_identifier = (x_user_id or "").strip()
    if not normalized_user_identifier or normalized_user_identifier == "anonymous":
        raise HTTPException(status_code=401, detail="X-User-Id header is required")
    user = resolve_authenticated_user(
        state,
        x_user_id=normalized_user_identifier,
        x_role=x_role,
    )
    if state.analytics_upload_store is None:
        return AnalyticsUploadHistoryResponse(
            items=[],
            store={"ready": False, "backend": "none"},
        )
    created_by = (
        None
        if user_has_permission(user, Permission.READ_ALL_ANALYTICS_UPLOADS)
        else user.user_identifier
    )
    items = [
        AnalyticsUploadHistoryItem.model_validate(item)
        for item in state.analytics_upload_store.list_uploads(
            limit=limit,
            created_by=created_by,
        )
    ]
    record_operation(
        state,
        "analytics-table-upload-history-list",
        {
            "actor": user.user_identifier,
            "count": len(items),
            "limit": limit,
            "scope": "all" if created_by is None else "created-by",
        },
    )
    return AnalyticsUploadHistoryResponse(
        items=items,
        store={"ready": True, "backend": state.analytics_upload_store.__class__.__name__},
    )


def _file_extension(file_name: str) -> str:
    if "." not in file_name:
        return ""
    return file_name.rsplit(".", maxsplit=1)[-1].lower()


def _read_csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    return _drop_empty_rows([[cell.strip() for cell in row] for row in reader])


def _read_workbook_rows(content: bytes) -> tuple[str, list[list[str]]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = _drop_empty_rows(
                [
                    [_cell_to_string(value) for value in row]
                    for row in worksheet.iter_rows(values_only=True)
                ]
            )
            if rows:
                return worksheet.title, rows
    finally:
        workbook.close()

    raise HTTPException(status_code=422, detail="workbook does not contain tabular rows")


def _drop_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    return [row for row in rows if any(cell.strip() for cell in row)]


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_response(
    *,
    file_name: str,
    size_bytes: int,
    extension: str,
    rows: list[list[str]],
    sheet_name: str | None,
    analysis_case: AnalysisCase = "audit-data",
) -> TableUploadAnalysisResponse:
    columns = [cell or f"field_{index + 1}" for index, cell in enumerate(rows[0])]
    normalized_rows = [_normalize_row(row, len(columns)) for row in rows[1:]]
    duplicate_row_count = len(normalized_rows) - len({tuple(row) for row in normalized_rows})
    column_profiles = [
        _build_column_profile(column_name=column, column_index=index, rows=normalized_rows)
        for index, column in enumerate(columns)
    ]
    empty_cell_count = sum(1 for row in normalized_rows for cell in row if cell == "")
    high_empty_columns = [
        column
        for column in column_profiles
        if normalized_rows and column.empty_count / len(normalized_rows) >= 0.3
    ]
    duplicate_column_names = [column for column, count in Counter(columns).items() if count > 1]
    audit_signals = _build_audit_signals(columns)
    duplicate_charge_signals = [
        "金额/费用字段",
        "患者/就诊字段",
        "日期/时间字段",
        "项目/药品/目录字段",
    ]
    has_duplicate_charge_base = all(signal in audit_signals for signal in duplicate_charge_signals)

    quality_findings = [
        (
            f"识别到 {len(normalized_rows)} 行数据和 {len(columns)} 个字段。"
            if normalized_rows
            else "仅识别到表头，未发现可分析数据行。"
        ),
        (
            f"发现 {empty_cell_count} 个空值单元，需要确认是否为业务允许缺失。"
            if empty_cell_count > 0
            else "未发现空值单元。"
        ),
        (
            f"发现 {duplicate_row_count} 条完全重复行。"
            if duplicate_row_count > 0
            else "未发现完全重复行。"
        ),
        (
            f"存在重复字段名：{'、'.join(duplicate_column_names)}。"
            if duplicate_column_names
            else "字段名未发现重复。"
        ),
    ]
    recommendations = [
        (
            "重复收费核验字段基础完整，可按患者/就诊、项目、日期和金额形成初筛分组。"
            if has_duplicate_charge_base
            else (
                "重复收费核验字段不完整，"
                "需补齐患者/就诊、项目、日期和金额字段后再进入正式审计判断。"
            )
        ),
        (
            "已识别医保支付字段，可进一步核对支付范围、报销口径和目录限制条件。"
            if "医保支付字段" in audit_signals
            else "未识别医保支付字段，当前更适合做文件质量预检和通用异常线索整理。"
        ),
        (
            f"优先核对高空值字段：{'、'.join(column.name for column in high_empty_columns)}。"
            if high_empty_columns
            else "字段完整度未触发高空值预警。"
        ),
    ]
    source_label = f"{extension.upper()} 工作簿" if extension != "csv" else "CSV 文件"
    if sheet_name:
        source_label = f"{source_label}（sheet: {sheet_name}）"

    case_status, case_metrics, case_findings = _build_case_analysis(
        analysis_case=analysis_case,
        columns=columns,
        rows=normalized_rows,
        empty_cell_count=empty_cell_count,
        duplicate_row_count=duplicate_row_count,
        audit_signals=audit_signals,
    )

    return TableUploadAnalysisResponse(
        name=file_name,
        analysis_case=analysis_case,
        analysis_case_label=_analysis_case_label(analysis_case),
        case_status=case_status,
        case_metrics=case_metrics,
        case_findings=case_findings,
        size_kb=max(1, round(size_bytes / 1024)),
        extension=extension,
        status="parsed",
        sheet_name=sheet_name,
        columns=column_profiles,
        row_count=len(normalized_rows),
        empty_cell_count=empty_cell_count,
        duplicate_row_count=duplicate_row_count,
        message=f"后端已完成 {source_label}的字段画像。",
        quality_findings=quality_findings,
        audit_signals=audit_signals,
        recommendations=recommendations,
    )


def _analysis_summary(response: TableUploadAnalysisResponse) -> dict[str, object]:
    return {
        "status": response.status,
        "analysis_case": response.analysis_case,
        "analysis_case_label": response.analysis_case_label,
        "case_status": response.case_status,
        "case_metrics": [metric.model_dump(mode="json") for metric in response.case_metrics],
        "case_findings": list(response.case_findings),
        "sheet_name": response.sheet_name,
        "row_count": response.row_count,
        "column_count": len(response.columns),
        "empty_cell_count": response.empty_cell_count,
        "duplicate_row_count": response.duplicate_row_count,
        "quality_findings": list(response.quality_findings),
        "audit_signals": list(response.audit_signals),
        "recommendations": list(response.recommendations),
        "columns": [column.model_dump(mode="json") for column in response.columns],
    }


def _analysis_case_label(analysis_case: AnalysisCase) -> str:
    return "财务杜邦分析" if analysis_case == "dupont" else "审计数据分析"


def _build_case_analysis(
    *,
    analysis_case: AnalysisCase,
    columns: list[str],
    rows: list[list[str]],
    empty_cell_count: int,
    duplicate_row_count: int,
    audit_signals: list[str],
) -> tuple[CaseStatus, list[TableAnalysisMetric], list[str]]:
    if analysis_case == "dupont":
        return _build_dupont_analysis(columns=columns, rows=rows)

    row_count = len(rows)
    metrics = [
        TableAnalysisMetric(
            key="row_count",
            label="数据行数",
            value=row_count,
            display_value=f"{row_count} 行",
            status="available",
        ),
        TableAnalysisMetric(
            key="duplicate_row_count",
            label="完全重复记录",
            value=duplicate_row_count,
            display_value=f"{duplicate_row_count} 条",
            status="available",
        ),
        TableAnalysisMetric(
            key="empty_cell_count",
            label="空值单元",
            value=empty_cell_count,
            display_value=f"{empty_cell_count} 个",
            status="available",
        ),
        TableAnalysisMetric(
            key="audit_signal_count",
            label="可识别审计维度",
            value=len(audit_signals),
            display_value=f"{len(audit_signals)} 类",
            status="available",
        ),
    ]
    findings = [
        (
            f"发现 {duplicate_row_count} 条完全重复记录，建议优先核对重复收费或重复入账。"
            if duplicate_row_count > 0
            else "未发现完全重复记录；仍需结合业务主键继续检查近似重复。"
        ),
        (
            f"发现 {empty_cell_count} 个空值单元，需确认必填字段和业务允许缺失范围。"
            if empty_cell_count > 0
            else "当前样本未发现空值单元。"
        ),
        (
            f"已识别可用于核验的字段维度：{'、'.join(audit_signals)}。"
            if audit_signals
            else "未识别到常用审计字段，请核对表头或补充字段映射。"
        ),
    ]
    return ("completed" if rows else "needs-input"), metrics, findings


def _build_dupont_analysis(
    *,
    columns: list[str],
    rows: list[list[str]],
) -> tuple[CaseStatus, list[TableAnalysisMetric], list[str]]:
    required_columns: dict[str, tuple[str, tuple[str, ...]]] = {
        "net_profit": (
            "净利润",
            ("净利润", "税后净利润", "net_profit", "net profit"),
        ),
        "revenue": (
            "营业收入",
            ("营业收入", "营业总收入", "销售收入", "revenue", "sales"),
        ),
        "average_assets": (
            "平均总资产",
            (
                "平均总资产",
                "平均资产总额",
                "总资产平均余额",
                "average_total_assets",
                "avg_assets",
            ),
        ),
        "average_equity": (
            "平均净资产",
            (
                "平均净资产",
                "平均股东权益",
                "所有者权益平均余额",
                "average_equity",
                "avg_equity",
            ),
        ),
    }
    indexes = {
        key: _find_column_index(columns, aliases)
        for key, (_, aliases) in required_columns.items()
    }
    missing = [required_columns[key][0] for key, index in indexes.items() if index is None]
    if missing:
        return (
            "needs-input",
            _unavailable_dupont_metrics(),
            [
                f"缺少杜邦分析必需字段：{'、'.join(missing)}。",
                "请使用页面案例模板的字段名，或将现有列重命名后重新分析。",
            ],
        )

    resolved_indexes = {key: int(index) for key, index in indexes.items() if index is not None}
    selected_row: list[str] | None = None
    selected_values: dict[str, Decimal] | None = None
    for row in reversed(rows):
        values = {
            key: _decimal_from_cell(row[index] if index < len(row) else "")
            for key, index in resolved_indexes.items()
        }
        if all(value is not None for value in values.values()):
            concrete = {key: value for key, value in values.items() if value is not None}
            if (
                concrete["revenue"] != 0
                and concrete["average_assets"] != 0
                and concrete["average_equity"] != 0
            ):
                selected_row = row
                selected_values = concrete
                break

    if selected_values is None:
        return (
            "needs-input",
            _unavailable_dupont_metrics(),
            [
                "没有找到四项指标均为有效数值且分母非零的数据行。",
                "请检查营业收入、平均总资产和平均净资产是否为非零数值。",
            ],
        )

    net_profit_margin = selected_values["net_profit"] / selected_values["revenue"]
    total_asset_turnover = selected_values["revenue"] / selected_values["average_assets"]
    equity_multiplier = selected_values["average_assets"] / selected_values["average_equity"]
    return_on_equity = selected_values["net_profit"] / selected_values["average_equity"]
    metrics = [
        _dupont_metric(
            key="net_profit_margin",
            label="销售净利率",
            value=net_profit_margin,
            percentage=True,
            formula="净利润 ÷ 营业收入",
        ),
        _dupont_metric(
            key="total_asset_turnover",
            label="总资产周转率",
            value=total_asset_turnover,
            formula="营业收入 ÷ 平均总资产",
        ),
        _dupont_metric(
            key="equity_multiplier",
            label="权益乘数",
            value=equity_multiplier,
            formula="平均总资产 ÷ 平均净资产",
        ),
        _dupont_metric(
            key="return_on_equity",
            label="净资产收益率",
            value=return_on_equity,
            percentage=True,
            formula="销售净利率 × 总资产周转率 × 权益乘数",
        ),
    ]
    period_index = _find_column_index(
        columns,
        ("期间", "年度", "年份", "日期", "period", "year"),
    )
    period = (
        selected_row[period_index].strip()
        if selected_row is not None
        and period_index is not None
        and period_index < len(selected_row)
        and selected_row[period_index].strip()
        else "最近一条有效记录"
    )
    findings = [
        f"{period}的净资产收益率为 {_format_percent(return_on_equity)}。",
        (
            "本次结果由销售净利率、总资产周转率和权益乘数三项可复核指标分解得出，"
            "可分别检查盈利能力、资产使用效率和财务杠杆。"
        ),
        "这是确定性公式计算结果，未调用外部大模型；正式结论仍需结合会计政策和同比口径复核。",
    ]
    return "completed", metrics, findings


def _unavailable_dupont_metrics() -> list[TableAnalysisMetric]:
    return [
        TableAnalysisMetric(
            key=key,
            label=label,
            value=None,
            display_value="待补充数据",
            formula=formula,
            status="unavailable",
        )
        for key, label, formula in (
            ("net_profit_margin", "销售净利率", "净利润 ÷ 营业收入"),
            ("total_asset_turnover", "总资产周转率", "营业收入 ÷ 平均总资产"),
            ("equity_multiplier", "权益乘数", "平均总资产 ÷ 平均净资产"),
            (
                "return_on_equity",
                "净资产收益率",
                "销售净利率 × 总资产周转率 × 权益乘数",
            ),
        )
    ]


def _dupont_metric(
    *,
    key: str,
    label: str,
    value: Decimal,
    formula: str,
    percentage: bool = False,
) -> TableAnalysisMetric:
    return TableAnalysisMetric(
        key=key,
        label=label,
        value=float(round(value, 8)),
        display_value=_format_percent(value) if percentage else f"{value:.2f}",
        formula=formula,
        status="available",
    )


def _format_percent(value: Decimal) -> str:
    return f"{value * Decimal('100'):.2f}%"


def _find_column_index(columns: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized_aliases = {_normalized_column_name(alias) for alias in aliases}
    for index, column in enumerate(columns):
        if _normalized_column_name(column) in normalized_aliases:
            return index
    return None


def _normalized_column_name(value: str) -> str:
    return "".join(
        character.lower()
        for character in value.strip()
        if character not in " _-（）()"
    )


def _decimal_from_cell(value: str) -> Decimal | None:
    normalized = value.strip().replace(",", "")
    if not normalized:
        return None
    try:
        parsed = Decimal(normalized)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def _normalize_row(row: list[str], column_count: int) -> list[str]:
    return [row[index].strip() if index < len(row) else "" for index in range(column_count)]


def _build_column_profile(
    *,
    column_name: str,
    column_index: int,
    rows: list[list[str]],
) -> TableColumnProfile:
    values = [row[column_index] if column_index < len(row) else "" for row in rows]
    non_empty_values = [value for value in values if value]
    return TableColumnProfile(
        name=column_name,
        type=_infer_column_type(column_name, values),
        empty_count=len(values) - len(non_empty_values),
        unique_count=len(set(non_empty_values)),
        sample_values=list(dict.fromkeys(non_empty_values))[:3],
        audit_hint=_infer_audit_hint(column_name),
    )


def _infer_column_type(name: str, values: list[str]) -> ColumnType:
    non_empty_values = [value for value in values if value]
    normalized_name = name.lower()
    if not non_empty_values:
        return "空列"
    identifier_keywords = (
        "id",
        "编号",
        "编码",
        "身份证",
        "患者",
        "就诊",
        "病历",
        "patient",
        "visit",
        "code",
    )
    if _matches_any(normalized_name, identifier_keywords):
        return "标识"
    if _ratio(non_empty_values, _looks_numeric) >= 0.8:
        return "数值"
    if _ratio(non_empty_values, _looks_like_date) >= 0.8:
        return "日期"
    return "文本"


def _ratio(values: list[str], predicate: Callable[[str], bool]) -> float:
    matched = sum(1 for value in values if predicate(value))
    return matched / len(values)


def _looks_numeric(value: str) -> bool:
    try:
        Decimal(value.replace(",", ""))
    except InvalidOperation:
        return False
    return True


def _looks_like_date(value: str) -> bool:
    normalized = value.strip().replace("/", "-").replace(".", "-")
    if not normalized:
        return False
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        try:
            parsed_date = date.fromisoformat(normalized)
        except ValueError:
            return False
        return 1900 <= parsed_date.year <= 2100
    return 1900 <= parsed.year <= 2100


def _infer_audit_hint(name: str) -> str:
    if _matches_any(name, ("金额", "费用", "price", "amount", "cost", "fee", "charge", "total")):
        return "金额字段，可用于收费合规和异常金额核验"
    if _matches_any(name, ("患者", "病人", "patient", "姓名", "身份证", "就诊", "visit")):
        return "对象字段，可用于同人同次就诊聚合"
    if _matches_any(name, ("日期", "时间", "date", "time", "结算", "发生")):
        return "时间字段，可用于限定审计期间和同日重复核验"
    if _matches_any(name, ("项目", "药品", "目录", "item", "drug", "catalog", "编码", "code")):
        return "项目字段，可用于目录限制和重复收费核验"
    if _matches_any(name, ("科室", "department", "dept")):
        return "组织字段，可用于科室维度分布分析"
    if _matches_any(name, ("医保", "结算", "支付", "报销", "insurance", "fund")):
        return "医保字段，可用于支付范围和报销口径核验"
    if _matches_any(name, ("数量", "qty", "quantity", "num")):
        return "数量字段，可用于数量异常和金额复算"
    return "通用字段"


def _build_audit_signals(column_names: list[str]) -> list[str]:
    signal_rules = [
        ("金额/费用字段", ("金额", "费用", "price", "amount", "cost", "fee", "charge", "total")),
        ("患者/就诊字段", ("患者", "病人", "patient", "姓名", "身份证", "就诊", "visit")),
        ("日期/时间字段", ("日期", "时间", "date", "time", "结算", "发生")),
        ("项目/药品/目录字段", ("项目", "药品", "目录", "item", "drug", "catalog", "编码", "code")),
        ("医保支付字段", ("医保", "结算", "支付", "报销", "insurance", "fund")),
        ("数量字段", ("数量", "qty", "quantity", "num")),
    ]
    return [
        label
        for label, keywords in signal_rules
        if any(_matches_any(column_name, keywords) for column_name in column_names)
    ]


def _matches_any(value: str, keywords: tuple[str, ...]) -> bool:
    normalized_value = value.lower()
    return any(keyword.lower() in normalized_value for keyword in keywords)
