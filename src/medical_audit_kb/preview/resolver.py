from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader

TEXT_FILE_SUFFIXES = {".md", ".txt"}
PDF_SUFFIX = ".pdf"
XLSX_SUFFIX = ".xlsx"
DEFAULT_CONTEXT_LINES = 2


@dataclass(frozen=True, slots=True)
class HighlightRange:
    start: int
    end: int
    text: str


@dataclass(frozen=True, slots=True)
class PreviewResult:
    source_path: Path
    media_type: str
    preview_text: str
    locator: dict[str, object]
    highlights: tuple[HighlightRange, ...]
    page_number: int | None = None
    line_start: int | None = None
    line_end: int | None = None
    sheet_name: str | None = None
    row_number: int | None = None


class PreviewResolutionError(ValueError):
    pass


class PreviewResolver:
    def __init__(self, *, source_root: Path | str) -> None:
        self._source_root = Path(source_root)

    def resolve(
        self,
        locator: dict[str, object],
        *,
        citation_text: str | None = None,
    ) -> PreviewResult:
        source_path = self._resolve_source_path(locator)
        suffix = source_path.suffix.lower()
        if suffix in TEXT_FILE_SUFFIXES:
            return self._resolve_text(source_path, locator, citation_text=citation_text)
        if suffix == PDF_SUFFIX:
            return self._resolve_pdf(source_path, locator, citation_text=citation_text)
        if suffix == XLSX_SUFFIX:
            return self._resolve_xlsx(source_path, locator, citation_text=citation_text)
        raise PreviewResolutionError(f"unsupported preview file type: {suffix}")

    def _resolve_source_path(self, locator: dict[str, object]) -> Path:
        raw_source_path = locator.get("source_path")
        if not isinstance(raw_source_path, str) or not raw_source_path:
            raise PreviewResolutionError("locator must contain source_path")

        source_path = Path(raw_source_path)
        if not source_path.is_absolute():
            source_path = self._source_root / source_path

        if not source_path.exists():
            raise FileNotFoundError(f"source file not found: {source_path}")
        return source_path

    def _resolve_text(
        self,
        source_path: Path,
        locator: dict[str, object],
        *,
        citation_text: str | None,
    ) -> PreviewResult:
        lines = source_path.read_text(encoding="utf-8").splitlines()
        line_start = _int_or_none(locator.get("line_start"))
        line_end = _int_or_none(locator.get("line_end"))
        article_number = _str_or_none(locator.get("article_number"))

        if line_start is None and article_number is not None:
            line_start, line_end = _find_article_lines(lines, article_number)

        if line_start is None:
            line_start = 1
        if line_end is None:
            line_end = line_start

        safe_start = max(1, line_start - DEFAULT_CONTEXT_LINES)
        safe_end = min(len(lines), line_end + DEFAULT_CONTEXT_LINES)
        preview_text = "\n".join(lines[safe_start - 1 : safe_end])
        return PreviewResult(
            source_path=source_path,
            media_type=_media_type(source_path),
            preview_text=preview_text,
            locator=dict(locator),
            highlights=_highlight_ranges(preview_text, citation_text),
            line_start=safe_start,
            line_end=safe_end,
        )

    def _resolve_pdf(
        self,
        source_path: Path,
        locator: dict[str, object],
        *,
        citation_text: str | None,
    ) -> PreviewResult:
        page_number = _int_or_none(locator.get("page_number"))
        if page_number is None:
            raise PreviewResolutionError("pdf locator must contain page_number")

        reader = PdfReader(source_path)
        if page_number < 1 or page_number > len(reader.pages):
            raise PreviewResolutionError(f"pdf page out of range: {page_number}")

        preview_text = reader.pages[page_number - 1].extract_text() or ""
        return PreviewResult(
            source_path=source_path,
            media_type=_media_type(source_path),
            preview_text=preview_text,
            locator=dict(locator),
            highlights=_highlight_ranges(preview_text, citation_text),
            page_number=page_number,
        )

    def _resolve_xlsx(
        self,
        source_path: Path,
        locator: dict[str, object],
        *,
        citation_text: str | None,
    ) -> PreviewResult:
        sheet_name = _str_or_none(locator.get("sheet_name"))
        row_number = _int_or_none(locator.get("row_number"))
        if sheet_name is None or row_number is None:
            raise PreviewResolutionError("xlsx locator must contain sheet_name and row_number")

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise PreviewResolutionError(f"worksheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
            values = next(
                worksheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    values_only=True,
                ),
                None,
            )
        finally:
            workbook.close()

        if values is None:
            raise PreviewResolutionError(f"xlsx row not found: {row_number}")

        cells = tuple("" if value is None else str(value).strip() for value in values)
        preview_text = "\n".join(
            f"{index + 1}: {cell}"
            for index, cell in enumerate(cells)
            if cell
        )
        return PreviewResult(
            source_path=source_path,
            media_type=_media_type(source_path),
            preview_text=preview_text,
            locator=dict(locator),
            highlights=_highlight_ranges(preview_text, citation_text),
            sheet_name=sheet_name,
            row_number=row_number,
        )


def _find_article_lines(lines: list[str], article_number: str) -> tuple[int, int]:
    pattern = re.compile(rf"^\s*{re.escape(article_number)}(?:\s|　|$)")
    start_index: int | None = None
    for index, line in enumerate(lines, start=1):
        if pattern.match(line):
            start_index = index
            break

    if start_index is None:
        raise PreviewResolutionError(f"article not found: {article_number}")

    next_article_pattern = re.compile(r"^\s*第[一二三四五六七八九十百千万零〇\d]+条(?:\s|　|$)")
    end_index = len(lines)
    for index in range(start_index + 1, len(lines) + 1):
        if next_article_pattern.match(lines[index - 1]):
            end_index = index - 1
            break
    return start_index, end_index


def _highlight_ranges(text: str, citation_text: str | None) -> tuple[HighlightRange, ...]:
    if not citation_text:
        return ()

    normalized_citation = re.sub(r"\s+", " ", citation_text).strip()
    if not normalized_citation:
        return ()

    exact_ranges = _find_all_ranges(text, normalized_citation)
    if exact_ranges:
        return exact_ranges

    ranges: list[HighlightRange] = []
    for token in _highlight_tokens(normalized_citation):
        ranges.extend(_find_all_ranges(text, token))
    return tuple(_dedupe_ranges(ranges))


def _find_all_ranges(text: str, needle: str) -> tuple[HighlightRange, ...]:
    if not needle:
        return ()
    ranges: list[HighlightRange] = []
    start = 0
    lowered_text = text.lower()
    lowered_needle = needle.lower()
    while True:
        index = lowered_text.find(lowered_needle, start)
        if index < 0:
            break
        end = index + len(needle)
        ranges.append(HighlightRange(start=index, end=end, text=text[index:end]))
        start = end
    return tuple(ranges)


def _highlight_tokens(text: str) -> tuple[str, ...]:
    tokens = re.findall(r"[a-zA-Z0-9]+|[\u4e00-\u9fff]{2,}", text)
    return tuple(token for token in tokens if len(token) >= 2)


def _dedupe_ranges(ranges: list[HighlightRange]) -> tuple[HighlightRange, ...]:
    seen: set[tuple[int, int]] = set()
    deduped: list[HighlightRange] = []
    for item in sorted(ranges, key=lambda value: (value.start, value.end)):
        key = (item.start, item.end)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return tuple(deduped)


def _int_or_none(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return None


def _str_or_none(value: object) -> str | None:
    return value if isinstance(value, str) and value else None


def _media_type(source_path: Path) -> str:
    suffix = source_path.suffix.lower()
    if suffix == ".md":
        return "text/markdown"
    if suffix == ".txt":
        return "text/plain"
    if suffix == ".pdf":
        return "application/pdf"
    if suffix == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "application/octet-stream"
