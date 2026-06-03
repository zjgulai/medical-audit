from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from typing import Protocol
from uuid import UUID

from openpyxl import load_workbook
from pypdf import PdfReader

from medical_audit_kb.domain.constants import FileErrorType, FileQueueStatus
from medical_audit_kb.domain.schemas import FailedFileCreate
from medical_audit_kb.ingestion.inventory import PENDING_MEDIA_TYPES, SUPPORTED_MEDIA_TYPES

MIN_EXTRACTABLE_PDF_CHARS = 20
MIN_EXTRACTABLE_TEXT_CHARS = 1
DEFAULT_MAX_XLSX_COLUMNS = 256


class WorksheetLike(Protocol):
    @property
    def max_column(self) -> int | None: ...

    @property
    def title(self) -> str: ...

    def iter_rows(self, *, max_col: int, values_only: bool) -> Iterable[tuple[object, ...]]: ...


class ExtractionStatus(StrEnum):
    EXTRACTED = "extracted"
    PENDING = "pending"
    FAILED = "failed"


@dataclass(frozen=True, slots=True)
class TextSegment:
    text: str
    page_number: int | None = None
    line_start: int | None = None
    line_end: int | None = None


@dataclass(frozen=True, slots=True)
class TableRow:
    sheet_name: str
    row_number: int
    cells: tuple[str, ...]
    values_by_header: dict[str, str]


@dataclass(frozen=True, slots=True)
class ExtractionResult:
    path: Path
    status: ExtractionStatus
    media_type: str
    text_segments: tuple[TextSegment, ...] = ()
    table_rows: tuple[TableRow, ...] = ()
    error_type: FileErrorType | None = None
    error_summary: str | None = None

    @property
    def text(self) -> str:
        return "\n".join(segment.text for segment in self.text_segments if segment.text)


def extract_file(path: Path | str) -> ExtractionResult:
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    media_type = SUPPORTED_MEDIA_TYPES.get(
        suffix,
        PENDING_MEDIA_TYPES.get(suffix, "application/octet-stream"),
    )

    if suffix in {".md", ".txt"}:
        return _extract_plain_text(source_path, media_type)
    if suffix == ".pdf":
        return _extract_pdf(source_path, media_type)
    if suffix == ".xlsx":
        return _extract_xlsx(source_path, media_type)
    if suffix in PENDING_MEDIA_TYPES:
        return _pending_result(source_path, media_type, FileErrorType.UNSUPPORTED_TYPE)
    return _pending_result(source_path, media_type, FileErrorType.UNSUPPORTED_TYPE)


def failed_file_payload_from_result(
    result: ExtractionResult,
    *,
    source_package_version_id: UUID,
    source_document_id: UUID | None = None,
    relative_path: str | None = None,
) -> FailedFileCreate:
    if result.status != ExtractionStatus.FAILED or result.error_type is None:
        raise ValueError("only failed extraction results can be converted to FailedFileCreate")

    return FailedFileCreate(
        source_package_version_id=source_package_version_id,
        source_document_id=source_document_id,
        relative_path=relative_path or result.path.as_posix(),
        error_type=result.error_type,
        error_summary=result.error_summary or "extraction failed",
        status=FileQueueStatus.OPEN,
    )


def _extract_plain_text(path: Path, media_type: str) -> ExtractionResult:
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        return _failed_result(path, media_type, FileErrorType.EXTRACTION_FAILED, str(exc))
    except OSError as exc:
        return _failed_result(path, media_type, FileErrorType.EXTRACTION_FAILED, str(exc))

    line_count = len(text.splitlines()) or 1
    if len(text.strip()) < MIN_EXTRACTABLE_TEXT_CHARS:
        return _pending_result(
            path,
            media_type,
            FileErrorType.LOW_QUALITY_TEXT,
            "text file has no extractable content",
        )
    return ExtractionResult(
        path=path,
        status=ExtractionStatus.EXTRACTED,
        media_type=media_type,
        text_segments=(TextSegment(text=text, line_start=1, line_end=line_count),),
    )


def _extract_pdf(path: Path, media_type: str) -> ExtractionResult:
    try:
        reader = PdfReader(path)
        segments = tuple(
            TextSegment(text=page.extract_text() or "", page_number=index + 1)
            for index, page in enumerate(reader.pages)
        )
    except Exception as exc:  # pypdf raises multiple parser-specific exceptions.
        return _failed_result(path, media_type, FileErrorType.EXTRACTION_FAILED, str(exc))

    total_text = "\n".join(segment.text for segment in segments).strip()
    if len(total_text) < MIN_EXTRACTABLE_PDF_CHARS:
        return _pending_result(
            path,
            media_type,
            FileErrorType.LOW_QUALITY_TEXT,
            "pdf has too little extractable text; treat as scanned or image-only file",
        )

    return ExtractionResult(
        path=path,
        status=ExtractionStatus.EXTRACTED,
        media_type=media_type,
        text_segments=segments,
    )


def _extract_xlsx(path: Path, media_type: str) -> ExtractionResult:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl can raise zip, XML, and workbook parser errors.
        return _failed_result(path, media_type, FileErrorType.EXTRACTION_FAILED, str(exc))

    try:
        rows: list[TableRow] = []
        for worksheet in workbook.worksheets:
            rows.extend(_extract_worksheet_rows(worksheet))
    finally:
        workbook.close()

    if not rows:
        return _pending_result(
            path,
            media_type,
            FileErrorType.LOW_QUALITY_TEXT,
            "xlsx workbook has no extractable rows",
        )

    return ExtractionResult(
        path=path,
        status=ExtractionStatus.EXTRACTED,
        media_type=media_type,
        table_rows=tuple(rows),
    )


def _extract_worksheet_rows(worksheet: WorksheetLike) -> list[TableRow]:
    max_column = min(int(worksheet.max_column or 0), DEFAULT_MAX_XLSX_COLUMNS)
    if max_column <= 0:
        return []

    rows: list[TableRow] = []
    header: tuple[str, ...] = ()
    for row_number, raw_values in enumerate(
        worksheet.iter_rows(max_col=max_column, values_only=True),
        start=1,
    ):
        cells = _normalize_row(raw_values)
        if not cells:
            continue
        if not header:
            header = cells
        rows.append(
            TableRow(
                sheet_name=str(worksheet.title),
                row_number=row_number,
                cells=cells,
                values_by_header=_values_by_header(header, cells),
            )
        )
    return rows


def _normalize_row(raw_values: tuple[object, ...]) -> tuple[str, ...]:
    cells = tuple("" if value is None else str(value).strip() for value in raw_values)
    trimmed = list(cells)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return tuple(trimmed)


def _values_by_header(header: tuple[str, ...], cells: tuple[str, ...]) -> dict[str, str]:
    if not header:
        return {}
    return {
        column_name: cells[index] if index < len(cells) else ""
        for index, column_name in enumerate(header)
        if column_name
    }


def _pending_result(
    path: Path,
    media_type: str,
    error_type: FileErrorType,
    error_summary: str | None = None,
) -> ExtractionResult:
    return ExtractionResult(
        path=path,
        status=ExtractionStatus.PENDING,
        media_type=media_type,
        error_type=error_type,
        error_summary=error_summary or "file type requires a later processing stage",
    )


def _failed_result(
    path: Path,
    media_type: str,
    error_type: FileErrorType,
    error_summary: str,
) -> ExtractionResult:
    return ExtractionResult(
        path=path,
        status=ExtractionStatus.FAILED,
        media_type=media_type,
        error_type=error_type,
        error_summary=error_summary,
    )
